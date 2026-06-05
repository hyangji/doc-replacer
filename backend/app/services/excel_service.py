"""Excel file parsing service for batch replacement mappings.

Parses Excel files (.xlsx) containing replacement data and maps
them to HWPX table structures for targeted modifications.
"""

import re
from io import BytesIO
from typing import Any

import openpyxl

from app.services.hwp_service import HwpService


class ExcelServiceError(Exception):
    """Base exception for Excel service errors."""


class ExcelParseError(ExcelServiceError):
    """Raised when Excel file parsing fails."""


# Known header variations (Korean and English)
FIELD_NAME_HEADERS = {"필드명", "필드", "field_name", "field", "항목", "항목명"}
OLD_VALUE_HEADERS = {"원본", "원본값", "old_value", "old", "기존", "기존값", "찾을내용", "찾기"}
NEW_VALUE_HEADERS = {"수정", "수정값", "new_value", "new", "변경", "변경값", "바꿀내용", "바꾸기"}


class ExcelService:
    """Service for Excel-based batch replacement operations."""

    def __init__(self) -> None:
        self._hwp_service = HwpService()

    async def parse_excel(self, file_content: bytes) -> dict:
        """Parse an Excel file and extract sheet-level data.

        Returns: {
            "sheets": [{
                "name": str,
                "headers": list[str],
                "rows": list[list[str]],
            }]
        }
        """
        try:
            wb = openpyxl.load_workbook(
                BytesIO(file_content), read_only=True, data_only=True
            )
        except Exception as e:
            raise ExcelParseError(f"엑셀 파일을 열 수 없습니다: {e}")

        sheets = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                sheets.append({"name": ws.title, "headers": [], "rows": []})
                continue

            headers = [str(c) if c is not None else "" for c in rows[0]]
            data_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows[1:]
                if row and any(c is not None for c in row)
            ]
            sheets.append({
                "name": ws.title,
                "headers": headers,
                "rows": data_rows,
            })

        wb.close()

        if not sheets:
            raise ExcelParseError("엑셀 파일에 시트가 없습니다.")

        return {"sheets": sheets}

    async def create_mapping(
        self, excel_data: dict, hwpx_tables: list[dict]
    ) -> list[dict]:
        """Create mapping rules between Excel data and HWPX table cells.

        Maps Excel rows to HWPX table cells by matching header names
        and first-column values.

        Args:
            excel_data: Output from parse_excel().
            hwpx_tables: Output from HwpService.extract_tables().

        Returns:
            List of {
                "table_index": int,
                "row": int,
                "col": int,
                "old_value": str,
                "new_value": str,
                "field_name": str,
            }
        """
        mappings: list[dict] = []

        # Extract replacement pairs from the first sheet
        replacements = self._extract_replacements_from_sheets(excel_data)

        for r in replacements:
            old_val = r["old_value"]
            for table in hwpx_tables:
                for row_idx, row in enumerate(table.get("rows", [])):
                    for col_idx, cell_text in enumerate(row):
                        if old_val and old_val in cell_text:
                            mappings.append({
                                "table_index": table["index"],
                                "row": row_idx,
                                "col": col_idx,
                                "old_value": old_val,
                                "new_value": r["new_value"],
                                "field_name": r.get("field_name", ""),
                            })

        return mappings

    async def apply_replacements(
        self, file_data: bytes, mappings: list[dict]
    ) -> tuple[bytes, list[dict]]:
        """Apply replacement mappings to HWPX file data (in-memory).

        Args:
            file_data: HWPX file bytes.
            mappings: List of mapping dicts (from create_mapping or parse_replacement_excel).

        Returns:
            (modified_file_data, change_log)
            change_log: [{"field_name": str, "old_value": str, "new_value": str, "count": int}]
        """
        current_data = file_data
        change_log: list[dict] = []

        for m in mappings:
            old_val = m.get("old_value", "")
            new_val = m.get("new_value", "")
            if not old_val:
                continue

            new_data, count = self._hwp_service.replace_text(
                current_data, old_val, new_val
            )
            change_log.append({
                "field_name": m.get("field_name", ""),
                "old_value": old_val,
                "new_value": new_val,
                "count": count,
            })
            if count > 0:
                current_data = new_data

        return current_data, change_log

    async def preview_mappings(
        self, excel_data: dict, hwpx_tables: list[dict]
    ) -> list[dict]:
        """Generate preview of what will be replaced (without modifying files).

        Returns:
            List of {
                "field_name": str,
                "old_value": str,
                "new_value": str,
                "match_count": int,
                "matched_locations": [{"table_index": int, "row": int, "col": int}],
            }
        """
        replacements = self._extract_replacements_from_sheets(excel_data)
        preview: list[dict] = []

        for r in replacements:
            old_val = r["old_value"]
            locations: list[dict] = []

            for table in hwpx_tables:
                for row_idx, row in enumerate(table.get("rows", [])):
                    for col_idx, cell_text in enumerate(row):
                        if old_val and old_val in cell_text:
                            locations.append({
                                "table_index": table["index"],
                                "row": row_idx,
                                "col": col_idx,
                            })

            preview.append({
                "field_name": r.get("field_name", ""),
                "old_value": old_val,
                "new_value": r["new_value"],
                "match_count": len(locations),
                "matched_locations": locations,
            })

        return preview

    def _extract_replacements_from_sheets(self, excel_data: dict) -> list[dict]:
        """Extract replacement pairs from parsed Excel data."""
        sheets = excel_data.get("sheets", [])
        if not sheets:
            return []

        # Use first sheet
        sheet = sheets[0]
        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])

        col_map = _detect_columns_from_list(headers)

        if col_map is None:
            # Fallback: assume A=field_name, B=old_value, C=new_value
            if len(headers) >= 3:
                col_map = {"field_name": 0, "old_value": 1, "new_value": 2}
            elif len(headers) >= 2:
                col_map = {"old_value": 0, "new_value": 1}
            else:
                return []

        replacements: list[dict] = []
        for row_idx, row in enumerate(rows):
            old_val = _safe_get(row, col_map.get("old_value"))
            new_val = _safe_get(row, col_map.get("new_value"))
            if not old_val:
                continue

            field_name = _safe_get(row, col_map.get("field_name"))
            if not field_name:
                field_name = f"row_{row_idx + 2}"

            replacements.append({
                "field_name": field_name,
                "old_value": old_val,
                "new_value": new_val,
            })

        return replacements


# ── Module-level functions (backward compatibility) ──


def parse_replacement_excel(file_content: bytes) -> list[dict]:
    """Parse an Excel file and extract replacement mappings.

    Returns: List of {"field_name": str, "old_value": str, "new_value": str}
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_content), read_only=True, data_only=True
        )
    except Exception as e:
        raise ExcelParseError(f"엑셀 파일을 열 수 없습니다: {e}")

    ws = wb.active
    if ws is None:
        raise ExcelParseError("엑셀 파일에 시트가 없습니다.")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ExcelParseError("엑셀 파일이 비어 있습니다.")

    col_map = _detect_columns(rows[0])
    data_rows = rows[1:]

    if col_map is None:
        if len(rows[0]) < 2:
            raise ExcelParseError(
                "엑셀 파일에 최소 2개 이상의 열이 필요합니다 (원본, 수정)."
            )
        col_map = _default_col_map(len(rows[0]))
        data_rows = rows

    replacements = []
    for row_idx, row in enumerate(data_rows, start=2):
        if not row or all(cell is None for cell in row):
            continue

        old_val = _cell_str(row, col_map["old_value"])
        new_val = _cell_str(row, col_map["new_value"])

        if not old_val:
            continue

        field_name = _cell_str(row, col_map.get("field_name"))
        if not field_name:
            field_name = f"row_{row_idx}"

        replacements.append({
            "field_name": field_name,
            "old_value": old_val,
            "new_value": new_val,
        })

    if not replacements:
        raise ExcelParseError("엑셀 파일에서 유효한 교체 항목을 찾을 수 없습니다.")

    return replacements


def parse_excel_preview(file_content: bytes) -> dict:
    """Parse Excel file and return preview data."""
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_content), read_only=True, data_only=True
        )
    except Exception as e:
        raise ExcelParseError(f"엑셀 파일을 열 수 없습니다: {e}")

    ws = wb.active
    if ws is None:
        raise ExcelParseError("엑셀 파일에 시트가 없습니다.")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ExcelParseError("엑셀 파일이 비어 있습니다.")

    headers = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = [
        [str(c) if c is not None else "" for c in row]
        for row in rows[1:]
        if row and any(c is not None for c in row)
    ]

    replacements = parse_replacement_excel(file_content)

    return {
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
        "replacements": replacements,
    }


def map_excel_to_tables(
    replacements: list[dict],
    tables: list[dict],
) -> list[dict]:
    """Map Excel replacement data to HWPX table cells for targeted replacement.

    Args:
        replacements: list of {"field_name": str, "old_value": str, "new_value": str}
        tables: list of {"index": int, "rows": list[list[str]], "headers": list[str]}

    Returns:
        List of replacement dicts with matched_tables locations.
    """
    mapped = []
    for r in replacements:
        old_val = r["old_value"]
        matches = []
        for table in tables:
            for row_idx, row in enumerate(table.get("rows", [])):
                for col_idx, cell_text in enumerate(row):
                    if old_val and old_val in cell_text:
                        matches.append({
                            "table_index": table["index"],
                            "row": row_idx,
                            "col": col_idx,
                        })
        mapped.append({
            **r,
            "matched_tables": matches,
        })
    return mapped


def build_replacement_diff(
    replacements: list[dict],
    old_text: str,
    new_text: str,
) -> list[dict]:
    """Build a diff record for each replacement showing before/after."""
    diff_records = []
    for r in replacements:
        count = old_text.count(r["old_value"])
        diff_records.append({
            "field_name": r["field_name"],
            "old_value": r["old_value"],
            "new_value": r["new_value"],
            "found_count": count,
            "applied": count > 0,
        })
    return diff_records


# ── Internal helpers ──


def _detect_columns(header_row: tuple) -> dict | None:
    if not header_row:
        return None
    headers_lower = [str(h).strip().lower() if h else "" for h in header_row]
    return _match_headers(headers_lower)


def _detect_columns_from_list(headers: list[str]) -> dict | None:
    if not headers:
        return None
    headers_lower = [h.strip().lower() for h in headers]
    return _match_headers(headers_lower)


def _match_headers(headers_lower: list[str]) -> dict | None:
    col_map: dict = {}
    for idx, h in enumerate(headers_lower):
        if h in FIELD_NAME_HEADERS:
            col_map["field_name"] = idx
        elif h in OLD_VALUE_HEADERS:
            col_map["old_value"] = idx
        elif h in NEW_VALUE_HEADERS:
            col_map["new_value"] = idx

    if "old_value" not in col_map or "new_value" not in col_map:
        return None
    return col_map


def _default_col_map(num_cols: int) -> dict:
    if num_cols >= 3:
        return {"field_name": 0, "old_value": 1, "new_value": 2}
    return {"old_value": 0, "new_value": 1}


def _cell_str(row: tuple, col_idx: int | None) -> str:
    if col_idx is None or col_idx >= len(row):
        return ""
    val = row[col_idx]
    return str(val).strip() if val is not None else ""


def _safe_get(row: list, col_idx: int | None) -> str:
    if col_idx is None or col_idx >= len(row):
        return ""
    val = row[col_idx]
    return str(val).strip() if val else ""


# ── 대비표 엑셀 파서 (신규) ──
# 도시계획 고시문의 기정/변경후 대비표 엑셀(시트 5개)을 파싱해
# (기정값, 변경후값) 교체쌍을 추출한다.
# 기존 parse_replacement_excel / ExcelService는 절대 변경하지 않는다.


# 델타 표기 패턴: "감) 526", "증) 110" 등
_DELTA_PATTERN = re.compile(r"^(감\)|증\))", re.IGNORECASE)


def _is_delta(val: Any) -> bool:
    """델타(증감) 표기인지 판별한다."""
    if val is None:
        return False
    s = str(val).strip()
    return bool(_DELTA_PATTERN.match(s))


def _is_blank_or_dash(val: Any) -> bool:
    """빈값 또는 대시('-') 계열인지 판별한다."""
    if val is None:
        return True
    s = str(val).strip()
    return s in ("", "-", "–", "—", "None")


def _normalize_number(val: Any) -> str | None:
    """엑셀 수치를 HWP 천단위 콤마 형식으로 변환한다.

    예:
        184636   -> "184,636"
        267309   -> "267,309"
        69.1     -> "69.1"
        100      -> "100"
        '184636' -> "184,636"

    반환값이 None이면 숫자가 아니거나 변환 불가.
    """
    if val is None:
        return None
    # 이미 콤마가 포함된 문자열이면 정규화
    if isinstance(val, str):
        s = val.strip().replace(",", "").replace("\xa0", "").replace(" ", "")
        if not s or s in ("-", "–", "—"):
            return None
        try:
            numeric = float(s)
        except ValueError:
            return None
    elif isinstance(val, (int, float)):
        numeric = float(val)
    else:
        return None

    # 정수이면 정수형, 소수이면 소수부 유지
    if numeric == int(numeric):
        int_val = int(numeric)
        if int_val < 0:
            return None  # 음수는 사용하지 않음
        return f"{int_val:,}"
    else:
        # 소수: 정수부에만 천단위 콤마
        int_part = int(numeric)
        frac_str = f"{numeric:.10f}".rstrip("0").split(".")[1]
        if int_part >= 1000:
            return f"{int_part:,}.{frac_str}"
        return f"{int_part}.{frac_str}"


def _cell_raw(row: tuple | list, idx: int) -> Any:
    """행에서 idx번 셀 원시값을 반환한다 (범위 초과 시 None)."""
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _label_text(row: tuple | list, *col_indices: int) -> str:
    """여러 컬럼 중 첫 번째 비어있지 않은 텍스트를 반환한다."""
    for idx in col_indices:
        val = _cell_raw(row, idx)
        if val is not None:
            s = str(val).strip().replace("\xa0", " ")
            if s:
                return s
    return ""


def _normalize_cell_text(val: Any) -> str:
    """셀 텍스트를 정규화한다 (공백/\xa0 제거, 소문자화하지 않음)."""
    if val is None:
        return ""
    return str(val).strip().replace("\xa0", "").replace(" ", "")


def _find_header_row(
    all_rows: list[tuple],
    keywords: list[str],
    max_scan: int = 5,
) -> int | None:
    """상위 max_scan 행을 스캔해 keywords가 모두 등장하는 행 인덱스를 반환한다.

    공백 정규화 후 비교하므로 '기 정' → '기정' 도 매칭된다.
    """
    for row_idx, row in enumerate(all_rows[:max_scan]):
        row_texts = [_normalize_cell_text(c) for c in row]
        if all(any(kw in t for t in row_texts) for kw in keywords):
            return row_idx
    return None


def _find_col_by_keyword(row: tuple | list, keyword: str) -> int | None:
    """행에서 keyword를 포함하는 첫 번째 셀 인덱스를 반환한다.

    공백 정규화 후 비교하므로 '기 정' → '기정' 도 매칭된다.
    """
    for idx, val in enumerate(row):
        if val is not None and keyword in _normalize_cell_text(val):
            return idx
    return None


# ── 시트 타입별 파서 ──


def _parse_sheet_type_a(
    all_rows: list[tuple],
    sheet_name: str,
) -> list[dict]:
    """타입 A: 멀티행 헤더 + 기정/변경/변경후 컬럼이 명시된 시트.

    대상: '5.나 유치업종', '8.나 토지이용계획'
    헤더 구조:
      행(n-1): 대분류 (면적/구성비 등)
      행(n):   기정 / 변경 / 변경후 / 증감
    데이터행: 행(n+1)~끝

    반환: list of {"sheet", "field_name", "old_value", "new_value"}
    """
    # 헤더 행 탐지: '기정'과 '변경'이 같이 있는 행
    header_row_idx = _find_header_row(all_rows, ["기정", "변경"], max_scan=6)
    if header_row_idx is None:
        return []

    header_row = all_rows[header_row_idx]
    data_rows = all_rows[header_row_idx + 1:]

    # '기정' 컬럼들과 대응하는 '변경후'(또는 '변경') 컬럼 탐지
    # 헤더 셀을 순회하며 (기정_idx, 변경후_idx) 쌍 수집
    # 우선순위: 기정 → 같은 그룹 내 '변경후' 우선, 없으면 '변경' 사용
    col_pairs: list[tuple[int, int]] = []  # (old_col, new_col)
    i = 0
    while i < len(header_row):
        cell_norm = _normalize_cell_text(header_row[i])
        if "기정" in cell_norm:
            # 이후 셀에서 '변경후' 우선 탐색 (최대 6칸 이내)
            found_new_col: int | None = None
            found_change_col: int | None = None
            for j in range(i + 1, min(i + 7, len(header_row))):
                nt = _normalize_cell_text(header_row[j])
                if "변경후" in nt:
                    found_new_col = j
                    break
                elif nt == "변경" and found_change_col is None:
                    found_change_col = j
                elif "기정" in nt:
                    # 다음 기정 그룹 시작 전에 멈춤
                    break
            if found_new_col is not None:
                col_pairs.append((i, found_new_col))
            elif found_change_col is not None:
                col_pairs.append((i, found_change_col))
        i += 1

    if not col_pairs:
        return []

    # 라벨 컬럼: 헤더 행의 앞쪽에서 라벨이 들어갈 컬럼 후보
    # 보통 1~3번 컬럼이 구분 라벨
    label_cols = [
        idx for idx in range(min(col_pairs[0][0], 5))
        if header_row[idx] is None or str(header_row[idx]).strip() == ""
    ]
    # 라벨 컬럼이 없으면 0번 컬럼 기본
    if not label_cols:
        label_cols = [0, 1, 2]

    # 상위 헤더(대분류) 행
    upper_header_row = all_rows[header_row_idx - 1] if header_row_idx > 0 else None

    results: list[dict] = []
    last_label = ""

    for row in data_rows:
        if not any(c is not None for c in row):
            continue

        # 라벨 추출 (여러 컬럼 중 비어있지 않은 것)
        label = _label_text(row, *range(min(col_pairs[0][0], len(row))))
        if label:
            last_label = label

        for old_col, new_col in col_pairs:
            old_raw = _cell_raw(row, old_col)
            new_raw = _cell_raw(row, new_col)

            # 제외: 빈값/대시
            if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
                continue
            # 제외: 델타 표기
            if _is_delta(old_raw) or _is_delta(new_raw):
                continue

            old_fmt = _normalize_number(old_raw)
            new_fmt = _normalize_number(new_raw)

            if old_fmt is None or new_fmt is None:
                continue
            # 값 동일하면 교체 불필요
            if old_fmt == new_fmt:
                continue

            # 대분류 컬럼 구분 (면적/구성비)
            category = ""
            if upper_header_row is not None:
                # old_col이 속하는 상위 헤더 탐색 (병합셀 고려: 왼쪽으로 검색)
                for uc in range(old_col, -1, -1):
                    uv = _cell_raw(upper_header_row, uc)
                    if uv is not None:
                        cat_text = str(uv).strip().replace("\xa0", " ")
                        if cat_text:
                            category = cat_text
                            break

            field_name = last_label or f"데이터행"
            if category and category not in field_name:
                field_name = f"{last_label} ({category})" if last_label else category

            results.append({
                "sheet": sheet_name,
                "field_name": field_name,
                "old_value": old_fmt,
                "new_value": new_fmt,
            })

    return results


def _parse_sheet_type_b(
    all_rows: list[tuple],
    sheet_name: str,
) -> list[dict]:
    """타입 B: 시트 7 (토지이용현황) — 교차형 기정/변경 컬럼 구조.

    시트 7에는 교차표가 2개 들어있다(가. 지목별 현황 / 나. 소유자별 현황).
    각 하위표 헤더 구조:
      행n-1: None | 구 분 | 계 | None | 답 | None | ...   (칼럼 카테고리, 병합셀)
      행n:   None | None | 기정 | 변경 | 기정 | 변경 | ... (기정/변경 교차)
    데이터행: 행n+1 ~ 다음 하위표 직전

    각 셀의 완전한 맥락 = [하위표명] × [칼럼 카테고리] × [지표].
    field_name 포맷: "지목별 현황 · 답 · 면적(㎡)"

    쌍: 기정 컬럼과 바로 다음 변경 컬럼.
    """
    # 1) 기정/변경 교차 헤더 행들을 모두 찾는다 (하위표마다 1개씩).
    header_row_indices: list[int] = []
    for idx, row in enumerate(all_rows):
        texts = [_normalize_cell_text(c) for c in row]
        if any(t == "기정" for t in texts) and any(t == "변경" for t in texts):
            header_row_indices.append(idx)

    if not header_row_indices:
        return []

    results: list[dict] = []

    for blk_i, header_idx in enumerate(header_row_indices):
        header_row = all_rows[header_idx]

        # 기정/변경 교차 컬럼 쌍 탐지
        col_pairs: list[tuple[int, int]] = []
        i = 0
        while i < len(header_row):
            cell_text = str(header_row[i]).strip() if header_row[i] is not None else ""
            if cell_text == "기정":
                next_text = str(header_row[i + 1]).strip() if i + 1 < len(header_row) and header_row[i + 1] else ""
                if next_text == "변경":
                    col_pairs.append((i, i + 1))
                    i += 2
                    continue
            i += 1

        if not col_pairs:
            continue

        # 칼럼 카테고리 행(헤더 바로 위 행) — 병합셀 대비 왼쪽 검색용
        category_row = all_rows[header_idx - 1] if header_idx - 1 >= 0 else None

        # 하위표 제목(가./나.) — 헤더 위쪽으로 스캔해 가장 가까운 상위 라벨
        subtable_name = _find_subtable_title(all_rows, header_idx)

        # 데이터행 범위: 헤더 다음 ~ 다음 헤더의 카테고리 행 직전(또는 끝)
        if blk_i + 1 < len(header_row_indices):
            # 다음 하위표의 카테고리 행(=다음 헤더-1)까지 제외
            data_end = header_row_indices[blk_i + 1] - 1
        else:
            data_end = len(all_rows)
        data_rows = all_rows[header_idx + 1:data_end]

        for row in data_rows:
            if not any(c is not None for c in row):
                continue

            metric = _label_text(row, 1)

            for old_col, new_col in col_pairs:
                old_raw = _cell_raw(row, old_col)
                new_raw = _cell_raw(row, new_col)

                if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
                    continue
                if _is_delta(old_raw) or _is_delta(new_raw):
                    continue

                old_fmt = _normalize_number(old_raw)
                new_fmt = _normalize_number(new_raw)

                if old_fmt is None or new_fmt is None:
                    continue
                if old_fmt == new_fmt:
                    continue

                # 칼럼 카테고리(지목/소유자) — 기정 컬럼에서 왼쪽으로 검색(병합셀)
                category = _find_category_label(category_row, old_col)

                parts = [p for p in (subtable_name, category, metric) if p]
                field_name = " · ".join(parts) if parts else "데이터행"

                results.append({
                    "sheet": sheet_name,
                    "field_name": field_name,
                    "old_value": old_fmt,
                    "new_value": new_fmt,
                })

    return results


def _find_subtable_title(all_rows: list[tuple], header_idx: int) -> str:
    """기정/변경 헤더 행 위쪽으로 스캔해 가장 가까운 '가. ...'/'나. ...' 하위표 제목을 찾는다.

    접두 "가. "/"나. " 등은 떼고 본문("지목별 현황")만 반환한다.
    """
    for idx in range(header_idx - 1, -1, -1):
        row = all_rows[idx]
        for c in row:
            if c is None:
                continue
            text = str(c).strip().replace("\xa0", " ")
            m = re.match(r"^[가-힣]\.\s*(.+)$", text)
            if m:
                return m.group(1).strip()
    return ""


def _find_category_label(category_row: tuple | list | None, ref_col: int) -> str:
    """칼럼 카테고리 행에서 ref_col(기정 컬럼)이 속한 상위 헤더 라벨을 반환한다.

    병합셀로 인해 라벨이 ref_col 왼쪽 칸에만 있을 수 있어 왼쪽으로 검색한다.
    '구 분' 같은 머릿글은 무시한다.
    """
    if category_row is None:
        return ""
    for idx in range(min(ref_col, len(category_row) - 1), -1, -1):
        val = category_row[idx]
        if val is None:
            continue
        text = str(val).strip().replace("\xa0", " ")
        if not text:
            continue
        if _normalize_cell_text(val) == "구분":
            return ""
        return text
    return ""


def _parse_sheet_8ga(
    all_rows: list[tuple],
    sheet_name: str,
) -> list[dict]:
    """8.가 용도지역계획 전용 파서.

    실제 헤더(dump 확인):
      행0: ... | '면 적 (㎡)' | ... | 구성비 | 비고
      행1: ... | '기 정' | '변 경' | None | '변경후' | ...

    공백 포함 '기 정', '변 경', '변경후' 처리.
    기정 컬럼의 일부 데이터는 '267,309\xa0' 같은 문자열 — _normalize_number가 처리.
    """
    # 공백 정규화 후 헤더 행 탐지
    header_row_idx = _find_header_row(all_rows, ["기정", "변경후"], max_scan=5)
    if header_row_idx is None:
        header_row_idx = _find_header_row(all_rows, ["기정", "변경"], max_scan=5)
        if header_row_idx is None:
            return []

    header_row = all_rows[header_row_idx]

    # 기정 컬럼 탐지 (공백 정규화)
    old_col = _find_col_by_keyword(header_row, "기정")
    if old_col is None:
        return []

    # '변경후' 컬럼 탐지 — old_col 이후에서 탐색
    new_col: int | None = None
    for j in range(old_col + 1, len(header_row)):
        nt = _normalize_cell_text(header_row[j])
        if "변경후" in nt:
            new_col = j
            break
    if new_col is None:
        # 변경후 없으면 '변경' 사용
        for j in range(old_col + 1, len(header_row)):
            nt = _normalize_cell_text(header_row[j])
            if "변경" in nt:
                new_col = j
                break
    if new_col is None:
        return []

    results: list[dict] = []
    data_rows = all_rows[header_row_idx + 1:]

    for row in data_rows:
        if not any(c is not None for c in row):
            continue

        label = _label_text(row, *range(min(old_col, 5)))

        old_raw = _cell_raw(row, old_col)
        new_raw = _cell_raw(row, new_col)

        if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
            continue
        if _is_delta(old_raw) or _is_delta(new_raw):
            continue

        old_fmt = _normalize_number(old_raw)
        new_fmt = _normalize_number(new_raw)

        if old_fmt is None or new_fmt is None:
            continue
        if old_fmt == new_fmt:
            continue

        results.append({
            "sheet": sheet_name,
            "field_name": label or "데이터행",
            "old_value": old_fmt,
            "new_value": new_fmt,
        })

    return results


def _split_sections_8da(
    all_rows: list[tuple],
) -> list[dict]:
    """8.다 시트를 '○ ...' 마커 기준으로 하위표 구간으로 분할한다.

    반환: list of {"label": 마커텍스트, "start": 마커행+1, "end": 다음마커행(또는 끝)}
    마커 이전(머릿글 등) 행은 무시한다.
    """
    markers: list[tuple[int, str]] = []
    for row_idx, row in enumerate(all_rows):
        for c in row:
            if c is not None and "○" in str(c):
                text = str(c).strip().replace("\xa0", " ")
                markers.append((row_idx, text))
                break

    sections: list[dict] = []
    for m_idx, (row_idx, text) in enumerate(markers):
        next_row = markers[m_idx + 1][0] if m_idx + 1 < len(markers) else len(all_rows)
        sections.append({
            "label": text,
            "start": row_idx + 1,
            "end": next_row,
        })
    return sections


def _parse_8da_road_summary(
    rows: list[tuple],
    sheet_name: str,
    section_label: str,
) -> list[dict]:
    """8.다 '○ 도로 총괄표' 구간 파서 (기정/변경 두 행 분리, 데이터 col3~).

    행 패턴:
      행i:   [None, '합 계', '기정', '14\xa0', '2,533\xa0', ...]
      행i+1: [None, None,    '변경', 14,       2534,       ...]
    구분/기정 라벨은 2번 컬럼, 데이터는 3번 컬럼부터.
    (기존 _parse_sheet_8da 총괄표 로직 유지)
    """
    results: list[dict] = []
    i = 0
    while i < len(rows) - 1:
        row = rows[i]
        kijong_cell = _normalize_cell_text(_cell_raw(row, 2))
        if "기정" not in kijong_cell:
            i += 1
            continue

        next_row = rows[i + 1]
        byeong_cell = _normalize_cell_text(_cell_raw(next_row, 2))
        if "변경" not in byeong_cell:
            i += 1
            continue

        parent_label_raw = _cell_raw(row, 1)
        if parent_label_raw is None or _normalize_cell_text(parent_label_raw) == "":
            for back in range(i - 1, max(i - 4, -1), -1):
                pv = _cell_raw(rows[back], 1)
                if pv is not None and _normalize_cell_text(pv):
                    parent_label_raw = pv
                    break

        parent_text = (
            str(parent_label_raw).strip().replace("\xa0", " ")
            if parent_label_raw else "데이터행"
        )

        for col_idx in range(3, max(len(row), len(next_row))):
            old_raw = _cell_raw(row, col_idx)
            new_raw = _cell_raw(next_row, col_idx)

            if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
                continue
            if _is_delta(old_raw) or _is_delta(new_raw):
                continue

            old_fmt = _normalize_number(old_raw)
            new_fmt = _normalize_number(new_raw)

            if old_fmt is None or new_fmt is None:
                continue
            if old_fmt == new_fmt:
                continue

            results.append({
                "sheet": sheet_name,
                "field_name": f"{section_label} {parent_text}".strip(),
                "old_value": old_fmt,
                "new_value": new_fmt,
            })

        i += 2

    return results


def _parse_8da_park_green_summary(
    rows: list[tuple],
    sheet_name: str,
    section_label: str,
) -> list[dict]:
    """8.다 '○ 공원·녹지 총괄' 구간 파서 (한 행에 기정/변경후 나란히).

    헤더:
      [None,'구분','기정',None,'변경',None,'변경후',None]
      [None, None,'면적','개소','면적','개소','면적','개소']
    데이터행:
      [None, 라벨, 기정면적(col2), 기정개소(col3), 증감(col4), '-'(col5),
       변경후면적(col6), 변경후개소(col7)]
    → col2(기정면적) → col6(변경후면적) 페어 추출.
    """
    results: list[dict] = []
    for row in rows:
        label = _label_text(row, 1)
        if not label:
            continue
        norm_label = _normalize_cell_text(label)
        # 헤더/부헤더 행 제외
        if norm_label in ("구분", "") or "면적" in norm_label or "개소" in norm_label:
            continue

        old_raw = _cell_raw(row, 2)  # 기정 면적
        new_raw = _cell_raw(row, 6)  # 변경후 면적

        if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
            continue
        if _is_delta(old_raw) or _is_delta(new_raw):
            continue

        old_fmt = _normalize_number(old_raw)
        new_fmt = _normalize_number(new_raw)

        if old_fmt is None or new_fmt is None:
            continue
        if old_fmt == new_fmt:
            continue

        results.append({
            "sheet": sheet_name,
            "field_name": label.replace("\xa0", " ").strip(),
            "old_value": old_fmt,
            "new_value": new_fmt,
        })

    return results


def _parse_8da_decision_table(
    rows: list[tuple],
    sheet_name: str,
    section_label: str,
) -> list[dict]:
    """8.다 결정조서 구간 파서 (기정/변경 두 행 분리, 면적 컬럼 기준).

    마커 행 다음에 헤더(2행)가 오고, 노선/시설별로 col1에 '기정' 또는 '변경'
    라벨이 들어간다. 변경이 있는 항목만 '변경' 행이 뒤따른다.

    구조 (공원/녹지 결정조서):
      [None,'기정', 번호, '공원'/'녹지', 세분, 위치, 면적(col6), '국토교통부고시']
    면적 컬럼은 헤더에서 '면적'을 포함하는 셀로 자동 탐지한다.
    면적 컬럼이 없으면(도로 결정조서 등) 빈 결과를 반환한다.

    페어링 규칙: '변경' 행은 직전의 '기정' 행과 짝지어 비교한다.
    """
    # 면적 컬럼 탐지 (구간 상단 헤더 행 스캔)
    area_col: int | None = None
    for row in rows[:4]:
        for idx, c in enumerate(row):
            if c is not None and "면적" in _normalize_cell_text(c):
                area_col = idx
                break
        if area_col is not None:
            break

    if area_col is None:
        return []

    results: list[dict] = []
    pending: tuple | None = None  # 직전 '기정' 행

    for row in rows:
        marker = _normalize_cell_text(_cell_raw(row, 1))
        if "기정" in marker:
            pending = row
            continue
        if "변경" in marker and pending is not None:
            old_raw = _cell_raw(pending, area_col)
            new_raw = _cell_raw(row, area_col)
            pending = None

            if _is_blank_or_dash(old_raw) or _is_blank_or_dash(new_raw):
                continue
            if _is_delta(old_raw) or _is_delta(new_raw):
                continue

            old_fmt = _normalize_number(old_raw)
            new_fmt = _normalize_number(new_raw)

            if old_fmt is None or new_fmt is None:
                continue
            if old_fmt == new_fmt:
                continue

            # 라벨: 세분명 + 위치 (변경 행 기준, 동일하므로 어느 쪽이든 무방)
            parts: list[str] = []
            for col_idx in range(2, area_col):
                lv = _cell_raw(row, col_idx)
                if lv is not None:
                    s = str(lv).strip().replace("\xa0", " ")
                    if s and s not in ("공원", "녹지", "-"):
                        parts.append(s)
            field_name = " ".join(parts).strip() or "데이터행"

            results.append({
                "sheet": sheet_name,
                "field_name": field_name,
                "old_value": old_fmt,
                "new_value": new_fmt,
            })

    return results


def _parse_8da_section(
    rows: list[tuple],
    sheet_name: str,
    section_label: str,
) -> list[dict]:
    """8.다 하위표 한 구간을 라벨에 따라 적절한 파서로 처리한다."""
    norm = _normalize_cell_text(section_label).replace("○", "")
    if "총괄표" in norm:
        # 도로 총괄표: 기정/변경 두 행 분리
        return _parse_8da_road_summary(rows, sheet_name, section_label)
    if "공원" in norm and "녹지" in norm and "총괄" in norm:
        # 공원·녹지 총괄: 단일행 col2→col6
        return _parse_8da_park_green_summary(rows, sheet_name, section_label)
    if "결정조서" in norm:
        # 결정조서: 기정/변경 두 행 분리 + 면적 컬럼
        return _parse_8da_decision_table(rows, sheet_name, section_label)
    return []


def _parse_sheet_8da(
    all_rows: list[tuple],
    sheet_name: str,
) -> list[dict]:
    """8.다 주요기반시설계획 전용 파서 (하위표 5개 전부 처리).

    구간:
      ○ 도로 총괄표      → 기정/변경 두 행 분리 (col3~)
      ○ 도로 결정조서    → 면적 컬럼 없음 → skip
      ○ 공원·녹지 총괄   → 단일행 col2(기정면적)→col6(변경후면적)
      ○ 공원 결정조서    → 면적 col6, 기정/변경 두 행 분리
      ○ 녹지 결정조서    → 면적 col6, 기정/변경 두 행 분리
    """
    items, _ = _parse_sheet_8da_with_sections(all_rows, sheet_name)
    return items


def _parse_sheet_8da_with_sections(
    all_rows: list[tuple],
    sheet_name: str,
) -> tuple[list[dict], list[dict]]:
    """8.다 파서 + 구간별 메타데이터를 함께 반환한다.

    반환: (items, sections)
      sections: list of {"sheet","label","extracted_count","status"}
        status: 'parsed'(추출>0) / 'empty'(데이터 있으나 추출 0) /
                'skipped'(면적 없는 세부 조서 등)
    """
    section_specs = _split_sections_8da(all_rows)
    all_items: list[dict] = []
    sections_meta: list[dict] = []

    for spec in section_specs:
        rows = all_rows[spec["start"]:spec["end"]]
        label = spec["label"]
        norm = _normalize_cell_text(label).replace("○", "")

        items = _parse_8da_section(rows, sheet_name, label)
        all_items.extend(items)

        if items:
            status = "parsed"
        elif "결정조서" in norm and "공원" not in norm and "녹지" not in norm:
            # 도로 결정조서 등 면적 없는 세부 조서
            status = "skipped"
        else:
            status = "empty"

        sections_meta.append({
            "sheet": sheet_name,
            "label": label,
            "extracted_count": len(items),
            "status": status,
        })

    return all_items, sections_meta


# ── 공개 함수 ──


def parse_comparison_table_excel(file_content: bytes) -> list[dict]:
    """도시계획 고시문 대비표 엑셀에서 (기정값, 변경후값) 교체쌍을 추출한다.

    지원 시트 구조:
    - 타입 A: 멀티행 헤더 + 기정/변경후 컬럼 명시 (5.나, 8.나)
    - 타입 B: 교차 기정/변경 컬럼 구조 (7. 토지이용현황)
    - 8.가 전용: 변경값이 두 컬럼에 분리된 특수 구조
    - 8.다 전용: 기정/변경이 행으로 교차하는 구조

    반환:
        list of {
            "sheet": 시트명,
            "field_name": 행 구분 라벨,
            "old_value": "184,636" (HWP 표기 일치 콤마 포맷),
            "new_value": "184,110",
        }

    예외:
        ExcelParseError: 파일 열기 실패 또는 유효 교체쌍이 없을 때
    """
    result = _parse_comparison_core(file_content)
    if not result["items"]:
        raise ExcelParseError(
            "엑셀 대비표에서 유효한 교체쌍을 찾을 수 없습니다. "
            "파일 형식이 도시계획 고시문 대비표(기정/변경후 구조)인지 확인하세요."
        )
    return result["items"]


def parse_comparison_table_excel_v2(file_content: bytes) -> dict:
    """대비표 엑셀 파서 v2 — 교체쌍 + 구간 메타데이터를 함께 반환한다.

    반환:
        {
            "items": [ {"sheet","field_name","old_value","new_value"}, ... ],
            "sections": [
                {"sheet": 시트명, "label": 구간 라벨,
                 "extracted_count": int, "status": str},
                ...
            ],
        }
        status:
          'parsed'  - 데이터를 정상 추출(추출 수 > 0)
          'empty'   - 데이터는 있으나 추출 0건(확인 필요)
          'skipped' - 면적 없는 세부 조서 등 의도적으로 건너뜀

    8.다 시트는 하위표 5개가 각각 별도 구간으로 분리되어 보고된다.
    그 외 시트는 시트 전체가 1개 구간으로 보고된다.

    예외:
        ExcelParseError: 파일 열기 실패 시.
        (유효 교체쌍이 0건이어도 예외를 던지지 않고 sections로 알린다.)
    """
    return _parse_comparison_core(file_content)


def _parse_comparison_core(file_content: bytes) -> dict:
    """대비표 엑셀 파싱 핵심 로직 (items + sections 동시 산출)."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelParseError(f"엑셀 파일을 열 수 없습니다: {e}")

    all_results: list[dict] = []
    all_sections: list[dict] = []

    try:
        for ws in wb.worksheets:
            sheet_name = ws.title
            all_rows = list(ws.iter_rows(values_only=True))
            # 완전히 빈 행 제거 (끝부분 패딩 행)
            non_empty_rows = [r for r in all_rows if any(c is not None for c in r)]

            if not non_empty_rows:
                continue

            if "8" in sheet_name and "다" in sheet_name:
                # 8.다: 하위표 5개를 각각 별도 구간으로 처리
                items, sections = _parse_sheet_8da_with_sections(
                    non_empty_rows, sheet_name
                )
                all_results.extend(items)
                all_sections.extend(sections)
                continue
            elif "8" in sheet_name and "가" in sheet_name:
                items = _parse_sheet_8ga(non_empty_rows, sheet_name)
            elif "7" in sheet_name and "토지이용현황" in sheet_name:
                items = _parse_sheet_type_b(non_empty_rows, sheet_name)
            else:
                # 기본: 타입 A (5.나, 8.나 등)
                items = _parse_sheet_type_a(non_empty_rows, sheet_name)

            all_results.extend(items)
            # 8.다 외 시트는 시트 전체를 1개 구간으로 보고
            all_sections.append({
                "sheet": sheet_name,
                "label": sheet_name,
                "extracted_count": len(items),
                "status": "parsed" if items else "empty",
            })
    finally:
        wb.close()

    return {"items": all_results, "sections": all_sections}
