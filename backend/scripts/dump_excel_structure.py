"""엑셀 대비표 구조 덤프 스크립트.

실제 샘플 파일의 각 시트 헤더/행 배치를 출력해 파서 작성 전 컬럼 위치를 확인한다.
"""

import sys
import openpyxl

EXCEL_PATH = "C:/Users/rkdgi/OneDrive/바탕 화면/고시문 샘플.xlsx"


def dump_sheet(ws, max_rows: int = 20) -> None:
    print(f"\n{'='*70}")
    print(f"시트명: {ws.title!r}")
    print(f"{'='*70}")
    all_rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in all_rows if any(c is not None for c in r)]
    for row_idx, row in enumerate(non_empty, start=1):
        if row_idx > max_rows:
            print(f"  ... (이후 행 생략, 총 {len(non_empty)}행)")
            break
        cells = [repr(c) for c in row]
        print(f"  행 {row_idx:2d}: {cells}")

    max_col = max((len(r) for r in non_empty), default=0)
    print(f"  => 전체 비어있지않은 행: {len(non_empty)}, 최대 컬럼수: {max_col}")


def main():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f"엑셀 열기 실패: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"파일: {EXCEL_PATH}")
    print(f"시트 목록: {wb.sheetnames}")

    for ws in wb.worksheets:
        dump_sheet(ws, max_rows=20)

    wb.close()
    print("\n완료.")


if __name__ == "__main__":
    main()
